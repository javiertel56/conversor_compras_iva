import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Color, numbers

def obtener_nombre(df, i):
    """
    Obtiene el nombre que está debajo del movimiento principal y arriba del IVA.
    """
    row = df.iloc[i]
    nombre = ""
    nombre_encontrado = False
    j = i + 1
    while j < len(df):
        next_row = df.iloc[j]
        # Si encontramos una fila de IVA, paramos
        if pd.notna(next_row[2]) and isinstance(next_row[2], str) and (
            "IVA" in next_row[4] if len(next_row) > 4 and pd.notna(next_row[4]) else False
        ):
            break
        # Si la columna 3 (nombre) tiene valor y no es una fila de cuenta, lo tomamos como nombre
        if pd.notna(next_row[3]) and (pd.isna(next_row[2]) or not (isinstance(next_row[2], str) and '-' in next_row[2])):
            nombre = next_row[3]
            nombre_encontrado = True
            break
        j += 1
    if not nombre_encontrado:
        nombre = row[3] if pd.notna(row[3]) else ""
    return nombre

def obtener_valor_iva(df, i):
    """
    Busca la fila de IVA asociada y devuelve (valor_iva, tipo_iva: '8' o '16').
    """
    for k in range(1, 6):
        if i + k < len(df):
            fila_check = df.iloc[i + k]
            cuenta_check = str(fila_check[2]) if pd.notna(fila_check[2]) else ""
            concepto_check = str(fila_check[4]) if len(fila_check) > 4 and pd.notna(fila_check[4]) else ""
            valor_iva = limpiar_valor(fila_check[6]) if len(fila_check) > 6 else 0.0
            if cuenta_check.startswith("1104-"):
                if cuenta_check.endswith("-02") or "8%" in cuenta_check or "8%" in concepto_check:
                    return valor_iva, '8'
                if cuenta_check.endswith("-01") or "16%" in cuenta_check or "16%" in concepto_check:
                    return valor_iva, '16'
    return 0.0, '16'

def obtener_tipo_iva(df, i):
    """
    Busca en las siguientes filas si hay cuenta de IVA 8% (-02 o 8%) o 16% (-01 o 16%).
    Devuelve '8' o '16'.
    """
    for k in range(1, 6):
        if i + k < len(df):
            fila_check = df.iloc[i + k]
            cuenta_check = str(fila_check[2]) if pd.notna(fila_check[2]) else ""
            concepto_check = str(fila_check[4]) if len(fila_check) > 4 and pd.notna(fila_check[4]) else ""
            if cuenta_check.endswith("-02") or  "8%" in cuenta_check or "8%" in concepto_check:
                return '8'
            if cuenta_check.endswith("-01") or "16%" in cuenta_check or "16%" in concepto_check:
                return '16'
    return '16'


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Inventario Tiendas")
        self.geometry("700x500")
        self.configure(bg="#1c2c47")
        self.selected_file = None
        self.create_tabs()

    def create_tabs(self):
        style = ttk.Style(self)
        style.theme_use('clam')

        dark_bg = "#1c2c47"
        dark_fg = "#f5f6fa"
        tab_bg = "#1c2c47"
        tab_active = "#1c2c47"

        style.configure('TNotebook', background=dark_bg, borderwidth=0)
        style.configure('TNotebook.Tab', background=tab_bg, foreground=dark_fg, font=('Segoe UI', 12, 'bold'), padding=[10, 5])
        style.map('TNotebook.Tab',
                  background=[('selected', tab_active)],
                  foreground=[('selected', '#00b894')])

        style.configure('TFrame', background=dark_bg)
        style.configure('TLabel', background=dark_bg, foreground=dark_fg, font=('Segoe UI', 10))

        style.configure('Blue.TButton',
                        background='#2980b9',
                        foreground='white',
                        borderwidth=0,
                        font=('Segoe UI', 11, 'bold'),
                        padding=10,
                        relief='flat')
        style.map('Blue.TButton',
                  background=[('active', '#1c5980')],
                  foreground=[('active', 'white')])

        style.configure('Yellow.TButton',
                        background='#fdcb6e',
                        foreground='#1c2c47',
                        borderwidth=0,
                        font=('Segoe UI', 11, 'bold'),
                        padding=10,
                        relief='flat')
        style.map('Yellow.TButton',
                  background=[('active', '#e1b84b')],
                  foreground=[('active', '#1c2c47')])

        style.configure('Green.TButton',
                        background='#00b894',
                        foreground='white',
                        borderwidth=0,
                        font=('Segoe UI', 11, 'bold'),
                        padding=10,
                        relief='flat')
        style.map('Green.TButton',
                  background=[('active', '#00916e')],
                  foreground=[('active', 'white')])

        notebook = ttk.Notebook(self)
        notebook.pack(expand=1, fill='both', padx=20, pady=20)

        rm_frame = ttk.Frame(notebook)
        tco_frame = ttk.Frame(notebook)

        notebook.add(rm_frame, text='Rosa Marcela')
        notebook.add(tco_frame, text='Tcomunicamos')

        self.create_tab_content(rm_frame, "Tabulador iva - Rosa Marcela")
        self.create_tab_content(tco_frame, "Tabulador iva - Tcomunicamos")

    def create_tab_content(self, frame, titulo):
        lbl_titulo = ttk.Label(frame, text=titulo, background="#1c2c47", foreground="#f5f6fa", font=('Segoe UI', 16, 'bold'))
        lbl_titulo.pack(pady=(10, 20))

        btn_frame = tk.Frame(frame, bg="#1c2c47")
        btn_frame.pack(pady=10)

        lbl_file = ttk.Label(frame, text="Archivo seleccionado: Ninguno", background="#1c2c47", foreground="#f5f6fa", font=('Segoe UI', 10))
        lbl_file.pack(pady=(10, 20))

        def subir_archivo():
            file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
            if file_path:
                self.selected_file = file_path
                lbl_file.config(text=f"Archivo seleccionado: {file_path}")

        def procesar_excel():
            if not self.selected_file:
                messagebox.showwarning("Advertencia", "Primero selecciona un archivo.")
                return
            try:
                destino = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
                if not destino:
                    return
                # Detecta el tab activo por el título del frame
                tab_tipo = "tcomunicamos" if "Tcomunicamos" in titulo else "rosa marcela"
                convertir_excel(self.selected_file, destino, tab_tipo)
                self.archivo_procesado = destino  # Guarda la ruta del archivo procesado
                messagebox.showinfo("Éxito", "Archivo procesado y guardado correctamente.")
            except Exception as e:
                messagebox.showerror("Error", f"Ocurrió un error: {e}")

        def abrir_archivo():
            # Abre el archivo procesado, no el original
            if hasattr(self, 'archivo_procesado') and self.archivo_procesado:
                import os
                os.startfile(self.archivo_procesado)
            else:
                messagebox.showwarning("Advertencia", "Primero procesa y guarda un archivo.")

        btn_subir = ttk.Button(btn_frame, text="Subir archivo", style="Blue.TButton", command=subir_archivo)
        btn_subir.pack(side=tk.LEFT, padx=10)

        btn_procesar = ttk.Button(btn_frame, text="Procesar Excel", style="Yellow.TButton", command=procesar_excel)
        btn_procesar.pack(side=tk.LEFT, padx=10)

        btn_abrir = ttk.Button(btn_frame, text="Abrir archivo", style="Green.TButton", command=abrir_archivo)
        btn_abrir.pack(side=tk.LEFT, padx=10)

def limpiar_valor(valor):
    if pd.isna(valor):
        return 0.0
    try:
        return float(str(valor).replace(',', '').replace(' ', ''))
    except Exception:
        return 0.0

def ajustar_formato_excel(ruta_archivo):
    wb = openpyxl.load_workbook(ruta_archivo)
    azul = PatternFill(start_color="2980b9", end_color="2980b9", fill_type="solid")
    # blanco = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Ya no se usa
    for ws in wb.worksheets:
        # Ajustar ancho de columnas
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Encabezado azul, negrita y centrado
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.fill = azul

        # Congelar la primera fila
        ws.freeze_panes = "A2"

        # Formato numérico y pintar de blanco los ceros en columnas de dinero e IVA
        header = [cell.value for cell in ws[1]]
        cols_dinero = []
        for nombre in ['Cargo 16', 'Abono 16', 'Cargo 8', 'Abono 8']:
            if nombre in header:
                cols_dinero.append(header.index(nombre) + 1)
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            for idx in cols_dinero:
                cell = row[idx-1]
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                if cell.value == 0 or cell.value == 0.0:
                    cell.font = Font(color="FFFFFF")  # Solo texto blanco
                    # cell.fill = blanco  # Elimina o comenta esta línea

    wb.save(ruta_archivo)
def convertir_excel(origen, destino, tab_tipo):
    # Lee el archivo original
    df = pd.read_excel(origen, header=None)
    movimientos = []
    hoja3 = []

    def obtener_tipo_iva_rosa_marcela(df, i):
        for k in range(1, 6):
            if i + k < len(df):
                fila_check = df.iloc[i + k]
                cuenta_check = str(fila_check[2]) if pd.notna(fila_check[2]) else ""
                concepto_check = str(fila_check[4]) if len(fila_check) > 4 and pd.notna(fila_check[4]) else ""
                if cuenta_check.endswith("-03") or  "8%" in cuenta_check or "8%" in concepto_check:
                    return '8'
                if cuenta_check.endswith("-02") or "16%" in cuenta_check or "16%" in concepto_check:
                    return '16'
        return '16'

    def obtener_valor_iva_rosa_marcela(df, i):
        for k in range(1, 6):
            if i + k < len(df):
                fila_check = df.iloc[i + k]
                cuenta_check = str(fila_check[2]) if pd.notna(fila_check[2]) else ""
                concepto_check = str(fila_check[4]) if len(fila_check) > 4 and pd.notna(fila_check[4]) else ""
                valor_iva = limpiar_valor(fila_check[6]) if len(fila_check) > 6 else 0.0
                if cuenta_check.endswith("-03") or  "8%" in cuenta_check or "8%" in concepto_check:
                    return valor_iva, '8'
                if cuenta_check.endswith("-02") or "16%" in cuenta_check or "16%" in concepto_check:
                    return valor_iva, '16'
        return 0.0, '16'

    i = 0
    while i < len(df):
        row = df.iloc[i]
        # Detecta fila de movimiento principal (usando el criterio original para identificar filas con cuenta)
        if pd.notna(row[0]) and pd.notna(row[2]) and isinstance(row[2], str) and '-' in row[2]:
            numero = row[0]
            referencia = row[1]
            cuenta = row[2]
            nombre = obtener_nombre(df, i)
            if tab_tipo == "rosa marcela":
                valor_iva, tipo_iva = obtener_valor_iva_rosa_marcela(df, i)
            else:
                valor_iva, tipo_iva = obtener_valor_iva(df, i) # Obtiene el valor de IVA

            # NUEVO FILTRO PRINCIPAL: Solo procesar filas si se encontró un valor de IVA mayor a cero
            if valor_iva > 0.0:
                concepto = ""
                # Usamos el cargo de la fila actual
                cargo = limpiar_valor(row[6]) if len(row) > 6 else 0.0
                abono = limpiar_valor(row[7]) if len(row) > 7 else 0.0

                # Busca el concepto en la siguiente fila si la columna de cuenta está vacía
                if i + 1 < len(df):
                    next_row = df.iloc[i + 1]
                    if pd.isna(next_row[2]) and pd.notna(next_row[4]):
                        concepto = str(next_row[4])
                    else:
                        concepto = nombre
                else:
                    concepto = nombre

                # Aplicamos la lógica de clasificación de IVA basada en el tipo_iva encontrado
                cargo_16 = abono_16 = cargo_8 = abono_8 = 0.0

                if tipo_iva == '8':
                    cargo_8 = round(cargo, 2) # Usamos el cargo de la fila actual
                    abono_8 = round(valor_iva, 2) # Usamos el valor_iva encontrado
                else: # Assume 16% if not 8%
                    cargo_16 = round(cargo, 2) # Usamos el cargo de la fila actual
                    abono_16 = round(valor_iva, 2) # Usamos el valor_iva encontrado

                # Filtra movimientos no deseados (finiquito, nomina, etc.)
                referencia_str = str(referencia).lower() if referencia else ""
                concepto_str = str(concepto).lower() if concepto else ""
                nombre_str = str(nombre).lower() if nombre else ""

                if (
                    "finiquito" in referencia_str or
                    "nomina" in referencia_str or
                    "sueldos y salarios" in concepto_str or
                    "sueldos y salarios" in nombre_str or
                    "imss" in concepto_str or
                    "imss" in nombre_str or
                    "comprobante sin requisitos" in concepto_str or
                    "comprobante sin requisitos" in nombre_str
                ):
                    i += 1 # Move to the next row
                    continue # Skip this row

                fila = [
                    numero, referencia, cuenta, nombre,
                    cargo_16, abono_16, cargo_8, abono_8,
                    None, None  # Las fórmulas se agregan después
                ]

                # Decide a qué hoja agregar la fila, ahora que sabemos que tiene IVA
                # Agregar a hoja3 si numero es 2 y la cuenta NO inicia con '1104-'
                # Mantenemos la verificación de cargo/abono > 0 para hoja3 como estaba previamente.
                if (
                    numero == 2 and
                    not (isinstance(cuenta, str) and cuenta.startswith("1104-")) and
                    (cargo_16 > 0 or abono_16 > 0 or cargo_8 > 0 or abono_8 > 0)
                ):
                    hoja3.append(fila)
                # Agregar a movimientos si cumple los criterios de cuenta y tiene cargo > 0
                elif (
                    cargo > 0 and
                    isinstance(cuenta, str) and (
                        cuenta.startswith("1102-") or
                        cuenta.startswith("5100-") or
                        cuenta.startswith("5200-") or
                        cuenta.startswith("61000-") or
                        cuenta.startswith("1201-")
                    )
                ):
                    movimientos.append(fila)

        i += 1 # Move to the next row

    columnas = [
        'No.', 'Refer.', 'Cuenta', 'Nombre',
        'Cargo 16', 'Abono 16', 'Cargo 8', 'Abono 8',
        'Fórmula 16', 'Fórmula 8'
    ]
    with pd.ExcelWriter(destino, engine='openpyxl') as writer:
        df_mov = pd.DataFrame(movimientos, columns=columnas)
        df_mov.to_excel(writer, index=False, sheet_name='Sheet1')
        if hoja3:
            df_hoja3 = pd.DataFrame(hoja3, columns=columnas)
            df_hoja3.to_excel(writer, index=False, sheet_name='Sheet3')

    # Agregar fórmulas en Excel y formato de dos decimales
    wb = openpyxl.load_workbook(destino)
    # Definir colores
    gris_suave = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Gris claro
    pastel = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")      # Verde pastel suave

    for sheet_name in ['Sheet1', 'Sheet3']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            max_row = ws.max_row
            # Formato y fórmulas por fila
            for row in range(2, max_row + 1):
                # Fórmula 16: =E2*0.16-F2
                ws[f'I{row}'] = f"=E{row}*0.16-F{row}"
                ws[f'I{row}'].number_format = '0.00'
                ws[f'I{row}'].fill = gris_suave
                # Fórmula 8: =G2*0.08-H2
                ws[f'J{row}'] = f"=G{row}*0.08-H{row}"
                ws[f'J{row}'].number_format = '0.00'
                ws[f'J{row}'].fill = gris_suave
                # Formato de dos decimales y colores en columnas de dinero
                for col in ['E', 'F', 'G', 'H']:
                    cell = ws[f'{col}{row}']
                    cell.number_format = '0.00'
                    if cell.value == 0 or cell.value == "0.00" or cell.value == 0.0:
                        cell.font = Font(color="FFFFFF")  # blanco
                    elif cell.value is not None and cell.value > 0:
                        cell.fill = pastel

            # Agregar suma al final de cada columna de dinero
            suma_row = max_row + 1
            ws[f'D{suma_row}'] = "TOTAL"
            ws[f'D{suma_row}'].font = Font(bold=True)
            for idx, col in enumerate(['E', 'F', 'G', 'H'], start=5):
                ws[f'{chr(64+idx)}{suma_row}'] = f"=SUM({chr(64+idx)}2:{chr(64+idx)}{max_row})"
                ws[f'{chr(64+idx)}{suma_row}'].number_format = '0.00'

            # Pintar la fila de totales de fórmulas de gris suave
            ws[f'I{suma_row}'].fill = gris_suave
            ws[f'J{suma_row}'].fill = gris_suave

    wb.save(destino)
    ajustar_formato_excel(destino)


if __name__ == "__main__":
    app = App()
    app.mainloop()
